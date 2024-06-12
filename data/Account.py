import ast
import openpyxl
import json
from Enum import FileLocationEnum


class Evm:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Evm)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)
    # def EVMAddress():


class Taproot:
    address = ['']
    segwit_address = ['']
    private_key = ['']
    segwit_privite_key = ['']
    mnemonic = ['']
    tb1p_address = ['']
    tb1p_private = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Taproot)
        # sheet_name = wb.active
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.segwit_address.append(cell.value)
        for row in sheet.iter_rows(min_col=5, max_col=5):
            for cell in row:
                self.segwit_privite_key.append(cell.value)
        for row in sheet.iter_rows(min_col=6, max_col=6):
            for cell in row:
                self.tb1p_address.append(cell.value)
        for row in sheet.iter_rows(min_col=7, max_col=7):
            for cell in row:
                self.tb1p_private.append(cell.value)
    # def EVMAddress():


class Polkadot:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Polkadot)
        # sheet_name = wb.active
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)
    # def EVMAddress():


class Solana:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Solana)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)


class Stack:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Stack)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)


class Cosmos:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Cosmos)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)


class Sui:
    address = ['']
    private_key = ['']
    mnemonic = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Sui)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.mnemonic.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.private_key.append(cell.value)


class Joy_id:
    address = ['']
    invite_code = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Joy_id)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.invite_code.append(cell.value)
    # def EVMAddress():


class ZealyTask:
    task_id = []
    submit_data = []
    FOLLOW_TWITTER = []
    JOIN_DISCORD = []

    QUIZ_ANSWER = {}
    TEXT_ANSWER = {}
    URL_ANSWER = {}
    TWITTER_REACT = []

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.ZealyTask)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.task_id.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.submit_data.append(eval(cell.value))
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                if cell.value:
                    self.FOLLOW_TWITTER = eval(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                if cell.value:
                    self.JOIN_DISCORD = eval(cell.value)
        for row in sheet.iter_rows(min_col=5, max_col=5):
            for cell in row:
                if cell.value:
                    self.QUIZ_ANSWER.update(json.loads(cell.value.replace("'", '"')))

        for row in sheet.iter_rows(min_col=6, max_col=6):
            for cell in row:
                if cell.value:
                    self.TEXT_ANSWER.update(json.loads(cell.value.replace("'", '"')))
        for row in sheet.iter_rows(min_col=7, max_col=7):
            for cell in row:
                if cell.value:
                    self.URL_ANSWER.update(json.loads(cell.value.replace("'", '"')))
        for row in sheet.iter_rows(min_col=8, max_col=8):
            for cell in row:
                if cell.value:
                    self.TWITTER_REACT = eval(cell.value)


class Gmail:
    gmail_address = ['']
    gmail_password = ['']
    gmail_name = ['']
    gmail_F2A = ['']
    gmail_remark = ['']

    def __init__(self):
        self.path = FileLocationEnum.Gmail
        self.wb = openpyxl.load_workbook(self.path)
        sheet_name = "Sheet1"
        sheet = self.wb[sheet_name]
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.gmail_address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.gmail_password.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.gmail_name.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.gmail_F2A.append(cell.value)
        for row in sheet.iter_rows(min_col=5, max_col=5):
            for cell in row:
                self.gmail_remark.append(cell.value)

    def saveRemark(self, browserId, msg):
        ws = self.wb.active

        ws.cell(row=browserId, column=5, value=str(msg))
        self.wb.save(self.path)


class ShitMails:
    shitmails_address = ['']
    shitmails_password = ['']
    shitmails_name = ['']

    def __init__(self):
        self.path = FileLocationEnum.Shitmails
        try:
            self.wb = openpyxl.load_workbook(self.path)
        except:
            self.wb = openpyxl.Workbook()

        sheet = self.wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.shitmails_address.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.shitmails_password.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.shitmails_name.append(cell.value)



class Twitter:
    twitter_name = ['']
    twitter_password = ['']
    twitter_F2A = ['']
    twitter_onePass = ['']
    twitter_email = ['']
    twitter_email_psd = ['']
    twitter_auth_token = ['']
    twitter_update_time = ['']
    twitter_id = ['']
    twitter_auth_v1_cookies = [[]]
    twitter_auth_v2_cookies = [[]]

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Twitter)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.twitter_name.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.twitter_password.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.twitter_F2A.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.twitter_onePass.append(cell.value)
        for row in sheet.iter_rows(min_col=5, max_col=5):
            for cell in row:
                self.twitter_email.append(cell.value)
        for row in sheet.iter_rows(min_col=6, max_col=6):
            for cell in row:
                self.twitter_email_psd.append(cell.value)
        for row in sheet.iter_rows(min_col=7, max_col=7):
            for cell in row:
                self.twitter_auth_token.append(cell.value)
        for row in sheet.iter_rows(min_col=8, max_col=8):
            for cell in row:
                self.twitter_update_time.append(cell.value)
        for row in sheet.iter_rows(min_col=9, max_col=9):
            for cell in row:
                self.twitter_id.append(cell.value)
        for row in sheet.iter_rows(min_col=10, max_col=10):
            for cell in row:
                if cell.value == 'NEW':
                    self.twitter_auth_v1_cookies.append(['NEW'])
                elif cell.value is not None and cell.value != "":
                    self.twitter_auth_v1_cookies.append(ast.literal_eval(cell.value))
                else:
                    self.twitter_auth_v1_cookies.append([])
        for row in sheet.iter_rows(min_col=11, max_col=11):
            for cell in row:
                if cell.value == 'NEW':
                    self.twitter_auth_v2_cookies.append(['NEW'])
                elif cell.value is not None and cell.value != "":
                    self.twitter_auth_v2_cookies.append(ast.literal_eval(cell.value))
                else:
                    self.twitter_auth_v2_cookies.append([])


class Discord:
    discord_email = ['']
    discord_password = ['']
    discord_username = ['']
    discord_token = ['']
    discord_2FA = ['']
    mail_password = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Discord)
        sheet_name = "Sheet1"
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.discord_email.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.discord_password.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.discord_username.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.discord_2FA.append(cell.value)
        for row in sheet.iter_rows(min_col=5, max_col=5):
            for cell in row:
                self.discord_token.append(cell.value)
        for row in sheet.iter_rows(min_col=6, max_col=6):
            for cell in row:
                self.mail_password.append(cell.value)


class Galxe:
    galxe_id = ['']
    authorization = ['']
    expires_time = []

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Galxe)
        sheet_name = "Sheet"
        sheet = wb.active
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.galxe_id.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.authorization.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.expires_time.append(cell.value)


class Zealy:
    zealy_address = ['']
    zealy_cookie = ['']
    zealy_id = ['']
    null_cookie = []

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(FileLocationEnum.Zealy)
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.zealy_address.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.zealy_id.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                ck = cell.value.__str__()
                if ck != 'None':
                    self.zealy_cookie.append(json.loads(ck.replace("'", '"')))
                # self.zealy_cookie.append(cell.value)
                # self.zealy_cookie.append(json.dumps(cell.value))
        for i in range(len(self.zealy_cookie)):
            if self.zealy_cookie[i] == '{}':
                self.null_cookie.append(i)
        print(f'Null Cookies:\n{self.null_cookie}')


class battle:
    battle_access = ['']
    access_token = ['']
    refresh_token = ['']
    id_token = ['']
    user_id = ['']

    def __init__(self):
        import openpyxl
        wb = openpyxl.load_workbook(r"C:\Users\crypto\Desktop\Account\battleshowdown.xlsx")
        sheet = wb.active
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                self.access_token.append(cell.value)
        for row in sheet.iter_rows(min_col=2, max_col=2):
            for cell in row:
                self.refresh_token.append(cell.value)
        for row in sheet.iter_rows(min_col=3, max_col=3):
            for cell in row:
                self.id_token.append(cell.value)
        for row in sheet.iter_rows(min_col=4, max_col=4):
            for cell in row:
                self.user_id.append(cell.value)
        # for row in sheet.iter_rows(min_col=2, max_col=2):
        #     for cell in row:
        #         self.battle_access.append(json.loads(cell.value))


if __name__ == '__main__':
    # pass
    evm = Evm()
    print(evm.address[20])
    # gmail = CreatName()
